"""Utilidades de importación y escritura de datos para ensayos mecánicos.

Expone lectura tolerante a encoding/formato (``import_data_text``,
``import_data_excel``), lectura/escritura de celdas individuales y en lotes
(``get_data_excel``, ``write_data_excel``, ``write_batch_excel``), y validación
de estructura de libros (``validate_excel_file``).
"""

from __future__ import annotations

import logging
import re
import pandas as pd
import numpy as np
import openpyxl as xl
import chardet
from pathlib import Path
from typing import Sequence, Tuple, Union, Any, Optional
from threading import Lock

# Configurar logging
logger = logging.getLogger(__name__)

# Opt-in to future pandas behavior to avoid FutureWarning about silent downcasting
try:
    pd.set_option('future.no_silent_downcasting', True)
except Exception:
    pass

__all__ = [
    'detect_encoding', 'import_data_text', 'import_data_excel', 'get_data_excel',
    'write_data_excel', 'write_batch_excel', 'write_multisheet_excel',
    'validate_excel_file',
]

_GLOBAL_CONVERSION_WARNING_EMITTED = False
_GLOBAL_CONVERSION_WARNING_LOCK = Lock()


def _should_emit_conversion_warning_once() -> bool:
    """Thread-safe gate to emit conversion warning only once globally."""
    global _GLOBAL_CONVERSION_WARNING_EMITTED
    with _GLOBAL_CONVERSION_WARNING_LOCK:
        if _GLOBAL_CONVERSION_WARNING_EMITTED:
            return False
        _GLOBAL_CONVERSION_WARNING_EMITTED = True
        return True


def detect_encoding(file_path: Union[str, Path], sample_size: int = 10000, default: str = 'utf-8') -> str:
    """Detecta la codificación probable de un archivo de texto.

    Args:
        file_path: Ruta al archivo.
        sample_size: Bytes a leer para muestreo (>0).
        default: Codificación a usar si no se detecta.

    Returns:
        Encoding detectado o la codificación por defecto.

    Raises:
        FileNotFoundError: Si el archivo no existe.
        ValueError: Si sample_size <= 0.
    """
    if sample_size <= 0:
        raise ValueError("sample_size debe ser mayor que 0")
    
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {file_path}")
    
    logger.debug(f"Detectando encoding de: {file_path}")
    
    try:
        with file_path.open('rb') as f:
            data = f.read(sample_size)
        result = chardet.detect(data) or {}
        encoding = result.get('encoding') or default
        confidence = result.get('confidence', 0.0)
        
        logger.debug(f"Encoding detectado: {encoding} (confianza: {confidence:.2f})")
        return encoding
    except OSError as e:
        logger.warning(f"Error detectando encoding, usando {default}: {e}")
        return default


_HEADER_KEYWORDS = ('running time', 'displacement')
_CANDIDATE_DELIMS = ('\t', ',', ';', '|')


def _read_preview_lines(file_path: Path, n_lines: int) -> list:
    """Lee hasta ``n_lines`` líneas del archivo para muestreo, con errors='ignore'."""
    try:
        with file_path.open('r', errors='ignore') as fh:
            out = []
            for _ in range(n_lines):
                try:
                    out.append(next(fh))
                except StopIteration:
                    break
            return out
    except OSError as e:
        logger.debug(f"No se pudo leer preview de {file_path}: {e}")
        return []


def _find_header_line(lines: Sequence[str]) -> Optional[int]:
    """Índice de la primera línea que contiene todas las keywords de encabezado."""
    for i, line in enumerate(lines):
        low = line.lower()
        if all(k in low for k in _HEADER_KEYWORDS):
            return i
    return None


def _detect_delimiter(lines: Sequence[str], expected_cols: int) -> Tuple[str, float]:
    """Devuelve ``(delimiter, score)`` del separador más probable entre candidatos.

    Prueba tab/coma/punto-y-coma/pipe y, si ninguno alcanza score ≥ 0.5,
    considera whitespace flexible (``r'\\s+'``).
    """
    sample = [ln for ln in lines if ln.strip()][:50]
    if not sample:
        return '\t', 0.0

    best_delim, best_score = '\t', -1.0
    for cand in _CANDIDATE_DELIMS:
        good = sum(1 for ln in sample if len(ln.strip().split(cand)) == expected_cols)
        score = good / len(sample)
        if score > best_score:
            best_delim, best_score = cand, score

    if best_score < 0.5:
        good = sum(1 for ln in sample if len(re.split(r'\s+', ln.strip())) >= expected_cols)
        score_ws = good / len(sample)
        if score_ws >= best_score:
            return r'\s+', score_ws

    return best_delim, best_score


def _clean_numeric_column(series: pd.Series) -> pd.Series:
    """Limpia texto ruidoso (espacios, NBSP, comas) y convierte a numérico con NaN."""
    cleaned = (
        series.astype(str)
        .str.strip()
        .str.replace('\u00a0', ' ', regex=False)
        .str.replace(',', '.', regex=False)
        .str.replace(r'[^0-9eE+\-\.]+', '', regex=True)
    )
    return pd.to_numeric(cleaned.replace('', np.nan), errors='coerce')


def _save_audit_sample(
    df: pd.DataFrame,
    audit_dir: Path,
    stem: str,
    head: int,
    mid: int,
    tail: int,
) -> None:
    """Guarda una muestra head/mid/tail del DataFrame como CSV para trazabilidad."""
    try:
        audit_dir.mkdir(parents=True, exist_ok=True)
        n = len(df)
        sample_idx = list(df.index[:head])
        if n > (head + tail):
            mid_start = max(0, n // 2 - mid // 2)
            sample_idx.extend(df.index[mid_start: mid_start + mid])
        sample_idx.extend(df.index[-tail:])
        sample_idx = sorted(set(sample_idx))
        out_path = audit_dir / f"{stem}.audit.csv"
        df.loc[sample_idx].to_csv(out_path, index=False)
        logger.debug(f"Auditoría guardada: {out_path}")
    except OSError as e:
        logger.debug(f"Auditoría no generada: {e}")


def import_data_text(
    file_path: Union[str, Path],
    delimiter: str,
    variable_names: Sequence[str],
    skiprows: int = 15,
    dtype: Any = np.float64,
    dropna: bool = True,
    max_retries: int = 2,
    coerce_numeric: bool = True,
    min_valid_cols: int = 2,
    debug_sample: bool = False,
    auto_detect_header: bool = True,
    warn_once: bool = True,
    audit_log_dir: Union[str, Path, None] = None,
    audit_head: int = 5,
    audit_tail: int = 5,
    audit_mid: int = 5,
) -> pd.DataFrame:
    """Importa datos de texto con estrategias robustas de limpieza y auditoría.

    - ``delimiter='auto'`` detecta el separador en un muestreo de las primeras líneas.
    - ``auto_detect_header`` localiza la línea 'Running Time … Displacement' y ajusta ``skiprows``.
    - ``coerce_numeric`` fuerza limpieza y conversión numérica columna por columna si la
      lectura directa con ``dtype`` falla.
    - ``min_valid_cols`` fija el mínimo de columnas no-NaN requerido para conservar una fila.
    - ``warn_once`` evita repetir el warning de conversión fallida entre hilos.
    - ``audit_log_dir`` — si se indica, guarda una muestra ``*.audit.csv`` (head/mid/tail).
    """
    if skiprows < 0:
        raise ValueError("skiprows debe ser >= 0")
    if not variable_names:
        raise ValueError("variable_names no puede estar vacío")

    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {file_path}")

    logger.info(f"Importando datos de texto: {file_path}")

    encodings_to_try = [detect_encoding(file_path), 'utf-8', 'latin-1', 'cp1252']

    if delimiter == 'auto':
        raw_lines = _read_preview_lines(file_path, n_lines=400)
        header_index = _find_header_line(raw_lines)
        data_section = raw_lines[(header_index + 2) if header_index is not None else 0:]
        delimiter, score = _detect_delimiter(data_section, len(variable_names))
        logger.info(f"Delimitador auto-detectado: {delimiter!r} (score={score:.2f})")

    if auto_detect_header:
        header_line = _find_header_line(_read_preview_lines(file_path, n_lines=200))
        if header_line is not None:
            new_skip = header_line + 2  # línea de títulos + línea de unidades
            if new_skip != skiprows:
                logger.info(f"Auto-detección encabezado: skiprows {skiprows} -> {new_skip}")
                skiprows = new_skip
        else:
            logger.debug("No se detectó encabezado; se mantiene skiprows proporcionado")

    warned_conversion = False
    last_error: Optional[Exception] = None
    df: pd.DataFrame | None = None

    read_kwargs = dict(
        names=list(variable_names),
        usecols=range(len(variable_names)),
        sep=delimiter,
        skiprows=skiprows,
        engine='python',
        on_bad_lines='skip',
    )

    for attempt, encoding in enumerate(encodings_to_try[:max_retries + 1]):
        try:
            logger.debug(f"Intento {attempt+1} lectura (encoding={encoding})")
            try:
                df = pd.read_csv(str(file_path), dtype=dtype, encoding=encoding, **read_kwargs)
            except ValueError as ve:
                if not (coerce_numeric and 'Unable to convert column' in str(ve)):
                    raise
                col_problem = str(ve).split('column')[-1].strip()
                should_warn = (
                    (not warned_conversion)
                    and ((not warn_once) or _should_emit_conversion_warning_once())
                )
                if should_warn:
                    logger.warning(f"Fallo conversión directa ({col_problem}). Reintentando con limpieza flexible.")
                warned_conversion = True
                df = pd.read_csv(str(file_path), dtype=str, encoding=encoding, **read_kwargs)
                if debug_sample:
                    logger.debug("Muestra cruda:\n" + '\n'.join(df.head(5).astype(str).agg('\t'.join, axis=1)))
                for col in df.columns:
                    df[col] = _clean_numeric_column(df[col])
            logger.info(f"Datos importados: {df.shape[0]} filas, {df.shape[1]} columnas (encoding={encoding})")
            break
        except (UnicodeDecodeError, UnicodeError) as e:
            last_error = e
            if attempt == max_retries:
                logger.error(f"Error de encoding tras {attempt+1} intentos: {e}")
                raise RuntimeError(f"Error de encoding: {e}")
            logger.warning(f"Encoding fallido ({encoding}), probando siguiente")
            continue
        except (OSError, pd.errors.ParserError, ValueError) as e:
            last_error = e
            logger.error(f"Error inesperado importando datos: {e}")
            raise
    if df is None:
        raise RuntimeError(f"No se pudo importar el archivo. Último error: {last_error}")

    # Fallback: si todo el DataFrame quedó NaN, reintento con auto-detección de sep
    # delegada a pandas (sep=None, engine='python').
    if df.isna().all(axis=1).mean() == 1.0:
        logger.warning("Todas las filas NaN tras coerción; intentando autodetección de separador")
        try:
            flex_df = pd.read_csv(
                str(file_path),
                sep=None,
                names=list(variable_names),
                usecols=range(len(variable_names)),
                dtype=str,
                skiprows=skiprows,
                engine='python',
                encoding=encodings_to_try[0],
                on_bad_lines='skip',
            )
            for col in flex_df.columns:
                flex_df[col] = _clean_numeric_column(flex_df[col])
            if (~flex_df.isna().all(axis=1)).any():
                df = flex_df
                logger.info("Fallback flexible exitoso")
        except (OSError, pd.errors.ParserError, ValueError) as fe:
            logger.error(f"Fallback flexible falló: {fe}")

    if dropna:
        initial = len(df)
        df = df.dropna(how='all')
        df = df[df.notna().sum(axis=1) >= min_valid_cols]
        removed = initial - len(df)
        if removed > 0:
            retention = (len(df) / initial * 100.0) if initial else 0.0
            logger.info(
                f"Filtrado de filas: eliminadas {removed} (all-NaN o < {min_valid_cols} válidas) "
                f"| Retención: {retention:.2f}% ({len(df)}/{initial})"
            )

    if audit_log_dir and not df.empty:
        _save_audit_sample(df, Path(audit_log_dir), file_path.stem, audit_head, audit_mid, audit_tail)

    if df.empty:
        logger.warning("DataFrame resultante está vacío")
    return df


def import_data_excel(
    file_path: Union[str, Path],
    sheet_idx: Union[int, str],
    variable_names: Sequence[str],
    skiprows: int = 49,
    dtype: Any = np.float64,
    dropna: bool = True,
    auto_detect_header: bool = True,
    header_search_rows: int = 100,
) -> pd.DataFrame:
    """Importa datos desde una hoja de Excel.

    Args:
        file_path: Ruta al archivo Excel.
        sheet_idx: Índice (int) o nombre (str) de la hoja.
        variable_names: Nombres de columnas a asignar.
        skiprows: Filas a omitir al inicio.
        dtype: Tipo de datos para conversión.
        dropna: Eliminar filas con NaN.
    
    Returns:
        DataFrame con los datos importados.
    
    Example:
        >>> df = import_data_excel('data.xlsx', 0, ['Time', 'Force', 'Disp'])
    """
    if skiprows < 0:
        raise ValueError("skiprows debe ser >= 0")
    if not variable_names:
        raise ValueError("variable_names no puede estar vacío")
    
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {file_path}")
    
    logger.info(f"Importando datos de Excel: {file_path}, hoja: {sheet_idx}")

    # Auto detección de fila inicial basada en coincidencia de nombres o patrón numérico
    if auto_detect_header:
        try:
            wb = xl.load_workbook(filename=str(file_path), read_only=True, data_only=True)
            if isinstance(sheet_idx, int):
                ws = wb.worksheets[sheet_idx]
            else:
                ws = wb[sheet_idx]
            target = [v.lower() for v in variable_names]
            header_candidate = None
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=header_search_rows, values_only=True), start=1):
                values = [str(v).strip().lower() for v in row if v is not None]
                if not values:
                    continue
                match_count = sum(1 for t in target if t in values)
                if match_count >= max(2, int(0.6 * len(target))):
                    # Asumimos siguiente fila son datos -> skip esa fila
                    header_candidate = r_idx
                    break
            if header_candidate is not None:
                new_skip = header_candidate
                if new_skip != skiprows:
                    logger.info(f"Auto-detección header Excel: skiprows {skiprows} -> {new_skip}")
                    skiprows = new_skip
            wb.close()
        except Exception as e:
            logger.debug(f"No se pudo auto-detectar header Excel: {e}")

    try:
        df = pd.read_excel(
            io=str(file_path),
            sheet_name=sheet_idx,
            names=list(variable_names),
            dtype=dtype,
            skiprows=skiprows,
            engine='openpyxl'
        )
        logger.info(f"Datos importados exitosamente: {df.shape[0]} filas, {df.shape[1]} columnas")
    except ValueError as e:
        if "Worksheet" in str(e):
            raise ValueError(f"La hoja '{sheet_idx}' no existe en el archivo Excel")
        raise
    except Exception as e:
        logger.error(f"Error importando Excel: {e}")
        raise
    
    if dropna:
        initial_rows = len(df)
        df = df.dropna()
        dropped_rows = initial_rows - len(df)
        if dropped_rows > 0:
            logger.info(f"Eliminadas {dropped_rows} filas con valores NaN")
    
    if df.empty:
        logger.warning("DataFrame resultante está vacío")
    
    return df


def get_data_excel(
    file_path: Union[str, Path],
    sheet_idx: Union[int, str, None],
    position: Tuple[int, int] = (1, 1)
) -> Any:
    """Obtiene un valor de una celda de un archivo Excel.

    Args:
        file_path: Ruta al archivo.
        sheet_idx: Índice (0-based) o nombre de la hoja. Si None usa la primera.
        position: (fila, columna) 1-based.
    
    Returns:
        Valor de la celda especificada.
    
    Example:
        >>> value = get_data_excel('data.xlsx', 'Results', (2, 3))
    """
    row, column = position
    if row < 1 or column < 1:
        raise ValueError("Las posiciones deben ser >= 1 (1-based indexing)")
    
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {file_path}")
    
    logger.debug(f"Leyendo celda ({row}, {column}) de {file_path}, hoja: {sheet_idx}")
    
    wb = xl.load_workbook(filename=str(file_path), keep_vba=True, read_only=True)
    try:
        if sheet_idx is None:
            sheet = wb.worksheets[0]
            logger.debug("Usando primera hoja del workbook")
        elif isinstance(sheet_idx, int):
            if sheet_idx >= len(wb.worksheets):
                raise ValueError(f"Índice de hoja {sheet_idx} fuera de rango (máximo: {len(wb.worksheets) - 1})")
            sheet = wb.worksheets[sheet_idx]
        else:
            if sheet_idx not in wb.sheetnames:
                raise ValueError(f"Hoja '{sheet_idx}' no encontrada. Hojas disponibles: {wb.sheetnames}")
            sheet = wb[sheet_idx]
        
        val = sheet.cell(row, column).value
        logger.debug(f"Valor leído: {val}")
        return val
        
    except Exception as e:
        logger.error(f"Error leyendo celda: {e}")
        raise
    finally:
        wb.close()


def _resolve_sheet(wb, sheet_name: str, create_sheet: bool):
    """Devuelve la hoja ``sheet_name``, creándola si no existe y ``create_sheet``."""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    if create_sheet:
        logger.info(f"Creada nueva hoja: {sheet_name}")
        return wb.create_sheet(title=sheet_name)
    raise ValueError(f"La hoja '{sheet_name}' no existe y create_sheet=False.")


def write_data_excel(
    file_path: Union[str, Path],
    sheet_name: str,
    position: Tuple[int, int],
    val: Any,
    create_sheet: bool = True
) -> None:
    """Escribe un valor en una celda de Excel.

    Args:
        file_path: Ruta al archivo existente.
        sheet_name: Nombre de hoja destino.
        position: (fila, columna) 1-based.
        val: Valor a escribir.
        create_sheet: Crear la hoja si no existe.
    """
    row, column = position
    if row < 1 or column < 1:
        raise ValueError("Las posiciones deben ser >= 1 (1-based indexing)")
    if not sheet_name.strip():
        raise ValueError("sheet_name no puede estar vacío")

    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {file_path}")

    logger.debug(f"Escribiendo valor {val} en ({row}, {column}) de hoja '{sheet_name}'")

    wb = xl.load_workbook(filename=str(file_path), keep_vba=True)
    try:
        sheet = _resolve_sheet(wb, sheet_name, create_sheet)
        sheet.cell(row, column, value=val)
        wb.save(str(file_path))
    finally:
        wb.close()


def write_batch_excel(
    file_path: Union[str, Path],
    sheet_name: str,
    data: Sequence[Tuple[int, int, Any]],
    create_sheet: bool = True
) -> None:
    """Escribe múltiples celdas en una misma hoja con una sola apertura del libro.

    Args:
        file_path: Ruta al archivo Excel.
        sheet_name: Nombre de la hoja destino.
        data: Secuencia de tuplas ``(fila, columna, valor)``.
        create_sheet: Crear la hoja si no existe.
    """
    if not data:
        logger.warning("Lista de datos vacía, no se escribirá nada")
        return

    if not sheet_name.strip():
        raise ValueError("sheet_name no puede estar vacío")

    for i, item in enumerate(data):
        if not isinstance(item, (tuple, list)) or len(item) != 3:
            raise ValueError(f"Elemento {i} debe ser tupla (fila, columna, valor)")
        row, col, _ = item
        if row < 1 or col < 1:
            raise ValueError(f"Elemento {i}: posiciones deben ser >= 1 (1-based)")

    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {file_path}")

    logger.info(f"Escribiendo {len(data)} valores en lote en hoja '{sheet_name}'")

    wb = xl.load_workbook(filename=str(file_path), keep_vba=True)
    try:
        sheet = _resolve_sheet(wb, sheet_name, create_sheet)
        for row, col, value in data:
            sheet.cell(row, col, value=value)
        wb.save(str(file_path))
    finally:
        wb.close()


def write_multisheet_excel(
    file_path: Union[str, Path],
    writes_by_sheet: "dict[str, Sequence[Tuple[int, int, Any]]]",
    create_sheet: bool = True,
) -> None:
    """Escribe múltiples celdas repartidas en varias hojas con una sola apertura.

    Args:
        file_path: Ruta al archivo Excel.
        writes_by_sheet: ``{sheet_name: [(row, col, value), ...]}``.
        create_sheet: Crear la hoja si no existe.
    """
    if not writes_by_sheet:
        logger.warning("writes_by_sheet vacío, no se escribirá nada")
        return

    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {file_path}")

    total = sum(len(v) for v in writes_by_sheet.values())
    logger.info(f"Escribiendo {total} celdas en {len(writes_by_sheet)} hoja(s)")

    wb = xl.load_workbook(filename=str(file_path), keep_vba=True)
    try:
        for sheet_name, cells in writes_by_sheet.items():
            if not cells:
                continue
            if not sheet_name.strip():
                raise ValueError("sheet_name no puede estar vacío")
            sheet = _resolve_sheet(wb, sheet_name, create_sheet)
            for row, col, value in cells:
                if row < 1 or col < 1:
                    raise ValueError(f"Posiciones deben ser >= 1 (1-based): ({row}, {col})")
                sheet.cell(row, col, value=value)
        wb.save(str(file_path))
    finally:
        wb.close()


def validate_excel_file(file_path: Union[str, Path]) -> dict:
    """Valida un archivo Excel y retorna información sobre sus hojas.
    
    Args:
        file_path: Ruta al archivo Excel.
    
    Returns:
        Diccionario con información del archivo: hojas, dimensiones, etc.
    
    Example:
        >>> info = validate_excel_file('data.xlsx')
        >>> print(info['sheets'])
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {file_path}")
    
    logger.info(f"Validando archivo Excel: {file_path}")
    
    wb = xl.load_workbook(filename=str(file_path), read_only=True)
    try:
        info = {
            'file_path': str(file_path),
            'file_size_mb': file_path.stat().st_size / (1024 * 1024),
            'sheets': [],
            'total_sheets': len(wb.worksheets)
        }
        
        for i, sheet in enumerate(wb.worksheets):
            sheet_info = {
                'index': i,
                'name': sheet.title,
                'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'dimensions': f"{sheet.max_row}x{sheet.max_column}"
            }
            info['sheets'].append(sheet_info)
        
        logger.info(f"Archivo validado: {info['total_sheets']} hojas")
        return info
        
    except Exception as e:
        logger.error(f"Error validando archivo: {e}")
        raise
    finally:
        wb.close()

